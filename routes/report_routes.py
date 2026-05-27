from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from utils.db import get_db
from utils.helpers import serialize
from bson import ObjectId
import datetime
import io

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

report_bp = Blueprint("reports", __name__)


def _get_report_data(db, site_id=None, start_date=None, end_date=None, report_type="material"):
    match = {}
    if site_id:
        match["site_id"] = site_id
    if start_date or end_date:
        match["date"] = {}
        if start_date:
            match["date"]["$gte"] = start_date
        if end_date:
            match["date"]["$lte"] = end_date

    # For equipment which uses created_at instead of date
    created_match = {}
    if site_id:
        created_match["site_id"] = site_id
    if start_date or end_date:
        created_match["created_at"] = {}
        if start_date:
            created_match["created_at"]["$gte"] = start_date
        if end_date:
            created_match["created_at"]["$lte"] = end_date

    rows = []

    if report_type in ["material", "all"]:
        logs = list(db.usage_logs.find(match).sort("date", -1))
        for log in logs:
            site = db.sites.find_one({"_id": ObjectId(log["site_id"])}) if ObjectId.is_valid(log.get("site_id", "")) else None
            mat = db.materials.find_one({"_id": ObjectId(log["material_id"])}) if ObjectId.is_valid(log.get("material_id", "")) else None
            cost = (mat["cost_per_unit"] if mat else 0) * log.get("used_quantity", 0)
            
            date_val = log.get("date")
            rows.append({
                "category": "Material",
                "date": date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime.datetime) else str(date_val),
                "site": site["name"] if site else "Unknown",
                "item": mat["name"] if mat else "Unknown",
                "unit": mat["unit"] if mat else "-",
                "quantity": log.get("used_quantity", 0),
                "cost_per_unit": mat["cost_per_unit"] if mat else 0,
                "total_cost": cost,
                "raw_date": date_val
            })

    if report_type in ["fuel", "all"]:
        logs = list(db.equipment_fuel.find(created_match).sort("created_at", -1))
        for log in logs:
            site = db.sites.find_one({"_id": ObjectId(log["site_id"])}) if ObjectId.is_valid(log.get("site_id", "")) else None
            eq = db.equipment.find_one({"_id": ObjectId(log["equipment_id"])}) if ObjectId.is_valid(log.get("equipment_id", "")) else None
            
            date_val = log.get("created_at")
            liters = log.get("liters", 0)
            total_cost = log.get("total_cost", 0)
            cpu = round(total_cost / liters, 2) if liters > 0 else 0
            
            rows.append({
                "category": "Fuel",
                "date": date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime.datetime) else str(date_val),
                "site": site["name"] if site else "Unknown",
                "item": eq["name"] if eq else "Unknown",
                "unit": "Liters",
                "quantity": liters,
                "cost_per_unit": cpu,
                "total_cost": total_cost,
                "raw_date": date_val
            })

    if report_type in ["maintenance", "all"]:
        logs = list(db.equipment_maintenance.find(created_match).sort("created_at", -1))
        for log in logs:
            site = db.sites.find_one({"_id": ObjectId(log["site_id"])}) if ObjectId.is_valid(log.get("site_id", "")) else None
            eq = db.equipment.find_one({"_id": ObjectId(log["equipment_id"])}) if ObjectId.is_valid(log.get("equipment_id", "")) else None
            
            date_val = log.get("created_at")
            cost = log.get("cost", 0)
            m_type = log.get("maintenance_type", "Repair")
            
            rows.append({
                "category": "Maintenance",
                "date": date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime.datetime) else str(date_val),
                "site": site["name"] if site else "Unknown",
                "item": f"{eq['name'] if eq else 'Unknown'} ({m_type})",
                "unit": "-",
                "quantity": 1,
                "cost_per_unit": cost,
                "total_cost": cost,
                "raw_date": date_val
            })

    if report_type == "all":
        rows.sort(key=lambda x: x["raw_date"] if x["raw_date"] else datetime.datetime.min, reverse=True)
        
    for r in rows:
        if "raw_date" in r:
            del r["raw_date"]

    return rows


@report_bp.route("/reports/usage", methods=["GET"])
@jwt_required()
def usage_report():
    db = get_db()
    site_id = request.args.get("site_id")
    start = request.args.get("start_date")
    end = request.args.get("end_date")
    report_type = request.args.get("type", "material")

    start_dt = datetime.datetime.fromisoformat(start) if start else None
    end_dt = datetime.datetime.fromisoformat(end) if end else None
    if end_dt:
        end_dt = end_dt.replace(hour=23, minute=59, second=59)

    rows = _get_report_data(db, site_id, start_dt, end_dt, report_type)
    total_cost = sum(r["total_cost"] for r in rows)
    return jsonify({"rows": rows, "total_cost": total_cost}), 200


@report_bp.route("/reports/export/pdf", methods=["GET"])
@jwt_required()
def export_pdf():
    if not REPORTLAB_AVAILABLE:
        return jsonify({"error": "reportlab not installed"}), 500

    db = get_db()
    site_id = request.args.get("site_id")
    start = request.args.get("start_date")
    end = request.args.get("end_date")
    report_type = request.args.get("type", "material")
    
    start_dt = datetime.datetime.fromisoformat(start) if start else None
    end_dt = datetime.datetime.fromisoformat(end) if end else None
    if end_dt:
        end_dt = end_dt.replace(hour=23, minute=59, second=59)

    rows = _get_report_data(db, site_id, start_dt, end_dt, report_type)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    title_map = {
        "material": "Material Usage Report",
        "fuel": "Equipment Fuel Report",
        "maintenance": "Equipment Maintenance Report",
        "all": "Comprehensive Expenditures Report"
    }
    title = title_map.get(report_type, "Expenditure Report")

    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Spacer(1, 12))

    headers = ["Date", "Site", "Category", "Item/Description", "Qty/Unit", "Cost/Unit", "Total Cost"]
    data = [headers]
    for row in rows:
        data.append([
            row["date"], row["site"], row["category"], row["item"],
            f"{row['quantity']} {row['unit']}", f"₹{row['cost_per_unit']:.2f}", f"₹{row['total_cost']:.2f}"
        ])

    total_cost = sum(r["total_cost"] for r in rows)
    data.append(["", "", "", "", "", "TOTAL", f"₹{total_cost:.2f}"])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a56db")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f3f4f6")]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fef3c7")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer, mimetype="application/pdf",
        as_attachment=True, download_name=f"{report_type}_report.pdf"
    )


@report_bp.route("/reports/export/excel", methods=["GET"])
@jwt_required()
def export_excel():
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "openpyxl not installed"}), 500

    db = get_db()
    site_id = request.args.get("site_id")
    start = request.args.get("start_date")
    end = request.args.get("end_date")
    report_type = request.args.get("type", "material")
    
    start_dt = datetime.datetime.fromisoformat(start) if start else None
    end_dt = datetime.datetime.fromisoformat(end) if end else None
    if end_dt:
        end_dt = end_dt.replace(hour=23, minute=59, second=59)

    rows = _get_report_data(db, site_id, start_dt, end_dt, report_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenditure Report"

    from openpyxl.styles import Font, PatternFill, Alignment
    headers = ["Date", "Site", "Category", "Item/Description", "Qty", "Unit", "Cost/Unit", "Total Cost"]
    header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row["date"])
        ws.cell(row=i, column=2, value=row["site"])
        ws.cell(row=i, column=3, value=row["category"])
        ws.cell(row=i, column=4, value=row["item"])
        ws.cell(row=i, column=5, value=row["quantity"])
        ws.cell(row=i, column=6, value=row["unit"])
        ws.cell(row=i, column=7, value=row["cost_per_unit"])
        ws.cell(row=i, column=8, value=row["total_cost"])

    total_cost = sum(r["total_cost"] for r in rows)
    last_row = len(rows) + 2
    ws.cell(row=last_row, column=7, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last_row, column=8, value=total_cost).font = Font(bold=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"{report_type}_report.xlsx"
    )
